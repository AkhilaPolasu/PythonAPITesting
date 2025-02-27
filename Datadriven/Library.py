import json
import jsonpath
import openpyxl
import requests
class Commom():
    def __init__(self,FileNamepath,SheetName):
        global wk
        global sh
        wk = openpyxl.load_workbook(FileNamepath)
        sh = wk[SheetName]

    def fetch_row_count(self):
        rows=sh.max_row
        return rows

    def fetch_column_count(self):
        columns= sh.max_column
        return columns

    def fetch_key_names(self):
        c=sh.max_column
        l=[]
        for i in range(1,c+1):
            cell=sh.cell(row=1,column=i)
            l.insert(i,cell.value)
        return l

    def update_request_with_data(self,rowNumber,jsonRequest,keyList):
        c=sh.max_column
        for i in range(1,c+1):
            cell= sh.cell(row=rowNumber,column=i)
            jsonRequest[keyList[i-1]]=cell.value
        return jsonRequest