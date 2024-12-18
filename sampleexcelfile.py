import json
import jsonpath
import requests
import openpyxl
def test_add_multiple_students():
    API_URL="http://thetestingworldapi.com/api/studentsDetails"
    f=open("/Inputfiles/Addnewmultiplstudent.json", "r")
    json_request = json.loads(f.read())
    print(json_request)
    wk=openpyxl.load_workbook("C:/Users/akhila.polasu/OneDrive - Calsoft Pvt Ltd/Desktop/EXELFiles/Book 6.xlsx")
    sh=wk['Sheet2']
    rows=sh.max_row
    print(rows)
    for i in range(2,rows+1):
        cell_first_name=sh.cell(row=i,column=1)
        cell_mid_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)
        json_request["first_name"]=cell_first_name.value
        json_request["middle_name"]=cell_mid_name.value
        json_request["last_name"]=cell_last_name.value
        json_request["date_of_birth"]=cell_dob.value
        response=requests.post(API_URL,json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201


#do not run it is stucking in terminal;